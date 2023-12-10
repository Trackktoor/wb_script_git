import openpyxl
from openpyxl import Workbook
import time
import traceback


class EXCEL_PARSER():
    def __init__(self, book_name='info.xlsx') -> None:
        self.book_name = book_name

    def get_work_book(self) -> Workbook:
        work_book = openpyxl.load_workbook(self.book_name).active
        return work_book

    def trip_el_excel(self,el):
        if el == None:
            return el
        if type(el) == int:
            return el
        return el.strip()

    def get_values(self) -> dict:
        work_book = self.get_work_book()

        start_row = 2
        end_row = work_book.max_row
        
        values = []

        for row in work_book.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            values.append(tuple(map(self.trip_el_excel, row)))
        return values

class EXCEL_REPORT():
    def __init__(self) -> None:
        self.count_product = 0

    def create_book(self):
        new_book = openpyxl.Workbook()
        page = new_book.active
        page.cell(row=1,column=1).value = '№'
        page.cell(row=1,column=2).value = 'Профиль'
        page.cell(row=1,column=3).value = 'Артикул'
        page.cell(row=1,column=4).value = 'Бренд'
        page.cell(row=1,column=5).value = 'Размер'
        page.cell(row=1,column=6).value = 'Текст запроса'
        page.cell(row=1,column=7).value = 'Статус'
        new_book.save('Отчет.xlsx')
        return new_book
    
    def get_target_book(self):
        while True:
            try:
                wb = openpyxl.load_workbook('Отчет.xlsx')
                return wb
            except:
                time.sleep(1)
                continue
    
    def add_product(self, info):
        while True:
            try:
                target_book = self.get_target_book()
                # self.count_product = EXCEL_PARSER(book_name='Отчет.xlsx').get_values()
                # print(self.count_product)
                # self.count_product = self.count_product[len(self.count_product)-1]
                # print(self.count_product)
                if len(info) == 7:
                    if info[4] == None: info[4] = ''

                    target_book.active.cell(target_book.active.max_row+1, column=1).value = 1
                    target_book.save('Отчет.xlsx')
                    target_book.active.cell(target_book.active.max_row, column=2).value = info[1]
                    target_book.active.cell(target_book.active.max_row, column=3).value = info[2]
                    target_book.active.cell(target_book.active.max_row, column=4).value = info[3]
                    target_book.active.cell(target_book.active.max_row, column=5).value = info[4]
                    target_book.active.cell(target_book.active.max_row, column=6).value = info[5]
                    target_book.active.cell(target_book.active.max_row, column=7).value = info[6]
                    target_book.save('Отчет.xlsx')
                    break
                else:
                    if info[4] == None: info[4] = ''

                    target_book.active.cell(target_book.active.max_row+1, column=1).value = 1
                    target_book.save('Отчет.xlsx')
                    target_book.active.cell(target_book.active.max_row, column=2).value = info[1]
                    target_book.active.cell(target_book.active.max_row, column=3).value = info[2]
                    target_book.active.cell(target_book.active.max_row, column=4).value = info[3]
                    target_book.active.cell(target_book.active.max_row, column=5).value = info[4]
                    target_book.active.cell(target_book.active.max_row, column=6).value = info[5]
                    target_book.save('Отчет.xlsx')
                    break
            except Exception as ex:
                time.sleep(2)
                print('Возможно нужно закрыть отчет')
                print(traceback.format_exc())
                continue
if __name__ == '__main__':
    # book = EXCEL_REPORT()
    # book.create_book()
    profile = 184811866
    parse = EXCEL_PARSER().get_values()
    parse = list(*filter(lambda info: info[1] == profile, parse))